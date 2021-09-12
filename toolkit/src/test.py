#import openpyxl

# Just a quick prototype
# just want to implement basic functions needed
# for the real deal

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Constant path to the spreadsheets
PATH_TO_SHEETS = "../test"

def save_workbook(wb, name):
    wb.save(f"{PATH_TO_SHEETS}/{name}")

def print_cell_value(ws, cell):
    print(ws[f'{cell}'].value)

def set_cell_value(ws, cell, val=""):
    ws[f'{cell}'].value = val

"""
 Will create a new sheet if the name is
 not already taken
"""
def create_new_sheet(wb, name):
    if name not in wb.sheetnames:
        wb.create_sheet(f"{PATH_TO_SHEETS}/{name}")
        #print(wb.sheetnames)
    else:
        print(f"{name} already exists!")

def main():
    # Get outselved a workbook
    wb = Workbook()

    # get the active worksheet
    selected_ws = wb.active

if __name__ == "__main__":
    main()