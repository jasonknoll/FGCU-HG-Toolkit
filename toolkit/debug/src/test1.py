# -*- coding: utf-8 -*-
"""
Created on Mon Sep 13 00:35:25 2021

@author: jason
"""

"""
 - TEST 1 - Intro 
 
 + First test: edit cells requested by player
 + Create, save, open, close workbooks
 + edit/remove/create sheets
 
 This is essentially just so I can practice working with excel until I 
 get access to the specific details of the project's process again.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter 

DEFAULT_SHEET_DIR = "../test1/"


def create_new_workbook(title=""):
    new_wb = Workbook()
    
    # TODO check if a title has been set
    # if not, it'll get some default name
    # that I'll set later
    
    if title != "":
        new_wb.title = title
    
    return new_wb


""" 
 Handle the saving process when closing
 workbooks and check for existing files 
"""
def close_workbook():
    pass

# ?
def edit_workbook():
    pass

def main():
    cli = CLI()
    cli.start()
    while cli.isRunning:
        cli.run()
    
    
    


"""
 Creating a CLI to handle all of the testing
 and practice needed
"""
class CLI:
    def __init__(self):
        self.isRunning = False
        self.wb = Workbook()
        self.ws = self.wb.active
        
        self.commands = {'exit': ["exit", "quit", "q"],
                         'new_workbook': [""]}
        self.args = {}
        
        
    # All CLI handling stuff    
    def start(self):
        self.isRunning = True
        
    def stop(self):
        self.isRunning = False
    
    def run(self):
        i = self.get_input()
    
    def get_input(self):
        return input("> ")
    
    """ 
     Do all the input "parsing".
     I wouldn't really call it parsing 
     because the structure of a command is very strict
     and hardcoded
    """
    def check_input(self, i):
        cmd_list = i.split()
        
        # gotta check if their quitting or not
        if (len(i) == 1):
            if i[0] in self.commands['exit']:
                self.stop()
            else:
                pass


if __name__ == "__main__":
    main()