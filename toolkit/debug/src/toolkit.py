# -*- coding: utf-8 -*-
"""
Created on Sun Sep 12 17:30:35 2021

@author: jason

FGCU Hydrogeology research team toolkit! 

First tool: allow user (research team member) to automate
their data entry. Eventually leading to automated data
visualizations to be used in presentations.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import matplotlib.pyplot as plt
import numpy as np

# Maybe set some global vars?? 
# I need to know where the sheets are and stuff


"""
 Main function will be our main loop running everything
 including the gui
"""
def main():
    hg_workbook = Workbook()
    hg_worksheet = hg_workbook.active
 

if __name__ == '__main__':
    main()