# -*- coding: utf-8 -*-
"""
Created on Mon Sep 13 00:35:25 2021

@author: jason
"""

"""
 - TEST 1 - Intro 
 
 + First test: edit cells requested by player
 + Create, save, open, close workbooks
 + edit/remove/create sheets
 
 This is essentially just so I can practice working with excel until I 
 get access to the specific details of the project's process again.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter 

DEFAULT_SHEET_DIR = "../test1/"


def create_new_workbook(title=""):
    new_wb = Workbook()
    # set the title if the arg is set
    if title != "":
        new_wb.title = title
    
    return new_wb


""" 
 Handle the saving process when closing
 workbooks and check for existing files 
"""
def close_workbook():
    pass

def save_workbook(wb, name):
    wb.save(f"{DEFAULT_SHEET_DIR}/{name}.xlsx")

# ?
def edit_workbook():
    pass

def main():
    cli = CLI()
    cli.start()
    while cli.isRunning:
        cli.run()
    


"""
 Creating a CLI to handle all of the testing
 and practice needed
"""
class CLI:
    def __init__(self):
        self.isRunning = False
        self.wb = Workbook()
        self.ws = self.wb.active
        
        self.commands = {'exit': ["exit", "quit", "q"],
                         'new_workbook': [""]}
        self.args = {}
        
        
    # All CLI handling stuff    
    def start(self):
        self.isRunning = True
        
    def stop(self):
        print("exiting...")
        self.isRunning = False
    
    def run(self):
        i = self.get_input()
        self.check_input(i)
    
    def get_input(self):
        return input("> ")
    
    """ 
     Do all the input "parsing".
     I wouldn't really call it parsing 
     because the structure of a command is very strict
     and hardcoded
    """
    def check_input(self, i):
        cmd_line = i.split()
        
        # gotta check if their quitting or not
        if (len(cmd_line) == 1):
            if cmd_line[0] in self.commands['exit']:
                self.stop()
            else:
                print(f"Sorry, {i[0]} is not a recognized command!")
        else:
            pass
        """
         TODO check for commands and arguments
         with opening and closing sheets. I'll get 
         to editing with the GUI.
         1. Open wb/create new one
        """

if __name__ == "__main__":
    main()