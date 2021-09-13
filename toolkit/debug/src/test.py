#import openpyxl

# Just a quick prototype
# just want to implement basic functions needed
# for the real deal

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# purely for testing matplotlib
import matplotlib.pyplot as plt
import numpy as np

fig, ax = plt.subplots()  # Create a figure containing a single axes.
ax.plot([1, 2, 3, 4], [1, 4, 2, 3])  # Plot some data on the axes.

plt.show()

# Constant path to the spreadsheets
PATH_TO_SHEETS = "../test"

def save_workbook(wb, name):
    wb.save(f"{PATH_TO_SHEETS}/{name}")

def print_cell_value(ws, cell):
    print(ws[f'{cell}'].value) 

def set_cell_value(ws, cell, val=""):
    ws[f'{cell}'].value = val

"""
 Will create a new sheet if the name is
 not already taken
"""
def create_new_sheet(wb, name):
    if name not in wb.sheetnames:
        wb.create_sheet(f"{PATH_TO_SHEETS}/{name}")
        #print(wb.sheetnames)
    else:
        print(f"{name} already exists!")

"""
 Allow the user to automatically add  
 a new entry to a spreadsheet that puts
 everything on a new line, it does the calculations
 correctly, and it'll edit other files if needed
"""

def add_sensor_entry():
    pass

def main():
    # Get outselved a workbook
    try:
        wb = load_workbook(f"{PATH_TO_SHEETS}/test0.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        print("File not found")

    # get the active worksheet
    selected_ws = wb.active
    # must pick 1 and 5 as excel is 1 indexed rather than 0
    for i in range(1, 5):
        print_cell_value(selected_ws, f"{get_column_letter(i)}1")

    pause_program()
    
def pause_program():
    input("press any key...")
    


if __name__ == "__main__":
    main()
