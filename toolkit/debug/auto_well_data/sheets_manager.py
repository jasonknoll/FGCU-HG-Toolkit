"""
 Manage the sheets' information
 including handling both old and new data,
 compiling data, and running the corrections on the
 depths. Afterward a new sheet will be generated
"""

from __future__ import print_function
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials

from openpyxl import Workbook, load_workbook

from tkinter import filedialog as fd
from tkinter import *

from dateutil import parser


class GoogleManager:
    def __init__(self):
        self.scopes = ['https://www.googleapis.com/auth/spreadsheets']
        self.old_sheet_id = '1RCAEGJuKwnAstDoXLw9eoRIZoJZMe2ZKNsTFyehBaVo'
        self.test_sheet_id = '1u8uMAEu6FZPHEoKERau1R_emPtARuAPBbDWfZZvY6Ao'
        
        """Shows basic usage of the Sheets API.
        Prints values from a sample spreadsheet.
        """
        creds = None
        # The file token.json stores the user's access and refresh tokens, and is
        # created automatically when the authorization flow completes for the first
        # time.
        if os.path.exists('googlesheets/token.json'):
            creds = Credentials.from_authorized_user_file('googlesheets/token.json', self.scopes)
        # If there are no (valid) credentials available, let the user log in.
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    'googlesheets/credentials.json', self.scopes)
                creds = flow.run_local_server(port=0)
            # Save the credentials for the next run
            with open('googlesheets/token.json', 'w') as token:
                token.write(creds.to_json())
    
        self.service = build('sheets', 'v4', credentials=creds)

        # Call the Sheets API
        self.sheet = self.service.spreadsheets()
        
        self.well_names = ['22M', '3A', 'MA', 'MM', '5M', 
                           '33M', 'U1', '27M', '6A', '28M', 
                           '29M', '11M', '10M', '7M', '1A', 
                           '9M', '12M', '30M', '5A', '32M', 
                           '31M', '18M', '2A', '21M']
        
        self.sheet_ids = {'22M': 1433469438, '3A': 480038452, 'MA': 707076060,
                          'MM':667590206, '5M': 898518076,'33M': 1686866287, 
                          'U1': 120162556, '27M': 960853423, '6A': 1309453391,
                          '28M': 2001958210,'29M': 1190619554, '11M': 2115651106,
                          '10M': 1181736078, '7M': 1307550729, '1A': 949756171, 
                           '9M': 67766467, '12M': 634987115, '30M': 674108455,
                           '5A': 44526233, '32M': 2024305940,'31M':1711889157,
                           '18M': 620154557, '2A': 1467669751, '21M': 1894757020}
        
        # just setting it to the first sheet for now
        self.curr_sheet = self.well_names[0]
        
        """
        self.result = self.sheet.values().get(spreadsheetId=self.test_sheet_id,
                                         range="22M!A2:D13").execute()
        self.values = self.result.get('values', [])
        
        
        print(self.values[0][0]) # A2
        print(self.values[1][0]) # A3
        #print(self.values[20][0]) # A22
        """
        
        
    # Range formatted 'Sheetname!Cell1:Cell2'    
    def update_cell(self, cell_range, value):
        self.sheet.values().update(spreadsheetId=self.test_sheet_id, 
                                   range=cell_range, 
                                   body={'values':[[value]]},
                                   valueInputOption="USER_ENTERED").execute()
        print("updated")
    
    def get_next_empty_row_manual_table(self):
        result = self.sheet.values().get(spreadsheetId=self.test_sheet_id,
                                         range=f'{self.curr_sheet}!G1:G100').execute()
        values = result.get('values', [])
        return len(values)+1
    
    def insert_manual_data_into_row(self, date, time, measure, row):
        
        self.update_cell(f"{self.curr_sheet}!G{row}", str(date))
        self.update_cell(f"{self.curr_sheet}!H{row}", str(time))
        self.update_cell(f"{self.curr_sheet}!I{row}", measure)
        print(row)
        
        # set to correct column
        reqs = {'requests': [
            {
                "repeatCell": {
                    "range": {
                      "sheetId": self.sheet_ids[f'{self.curr_sheet}'],
                      "startRowIndex": 2,
                      "endRowIndex": row,
                      "startColumnIndex": 6,
                      "endColumnIndex": 7
                    },
                    "cell": {
                      "userEnteredFormat": {
                        "numberFormat": {
                          "type": "DATE",
                          "pattern": "mm/dd/yy"
                        }
                      }
                    },
                    "fields": "userEnteredFormat.numberFormat"
                  }
            }, {
                "repeatCell": {
                    "range": {
                      "sheetId": self.sheet_ids[f'{self.curr_sheet}'],
                      "startRowIndex": 2,
                      "endRowIndex": row,
                      "startColumnIndex": 7,
                      "endColumnIndex": 8
                    },
                    "cell": {
                      "userEnteredFormat": {
                        "numberFormat": {
                          "type": "DATE",
                          "pattern": "hh:mm am/pm"
                        }
                      }
                    },
                    "fields": "userEnteredFormat.numberFormat"
                  }
                },
                {
                  "repeatCell": {
                    "range": {
                      "sheetId": self.sheet_ids[f'{self.curr_sheet}'],
                      "startRowIndex": 2,
                      "endRowIndex": row,
                      "startColumnIndex": 8,
                      "endColumnIndex": 9
                    },
                    "cell": {
                      "userEnteredFormat": {
                        "numberFormat": {
                          "type": "NUMBER",
                          "pattern": "0.00"
                        }
                      }
                    },
                    "fields": "userEnteredFormat.numberFormat"
                  }
                },
                {
                  "repeatCell": {
                    "range": {
                      "sheetId": self.sheet_ids[f'{self.curr_sheet}'],
                      "startRowIndex": 2,
                      "endRowIndex": row,
                      "startColumnIndex": 2,
                      "endColumnIndex": 3
                    },
                    "cell": {
                      "userEnteredFormat": {
                        "numberFormat": {
                          "type": "NUMBER",
                          "pattern": "0.00"
                        }
                      }
                    },
                    "fields": "userEnteredFormat.numberFormat"
                  }
                }
            ]}
        
        self.sheet.batchUpdate(spreadsheetId=self.test_sheet_id,
                               body=reqs).execute()
        
    def insert_formula_into_results_table(self, row):
        self.update_cell(f'{self.curr_sheet}!A{row}', f'=G{row}')
        self.update_cell(f'{self.curr_sheet}!B{row}', f'=H{row}')
        
        self.update_cell(f'{self.curr_sheet}!C{row}', f'=D{row}/305')
        self.update_cell(f'{self.curr_sheet}!D{row}', f'=I{row}')
        
    def submit_entry(self, date, time, measure):
        date_val = parser.parse(date.get())
        time_val = parser.parse(time.get())
        measure_val = float(measure.get())
        
        next_row = self.get_next_empty_row_manual_table()
        #print(next_row)
        self.insert_manual_data_into_row(date_val, time_val, measure_val, next_row)
        self.insert_formula_into_results_table(next_row)
    
    def get_well_by_name(self, name):
        for well in self.well_names:
            if name == well:
                return well
    
    def setup_dropdown(self, frame):
        
        var = StringVar(frame)
        var.set(self.well_names[0])
        dropdown = OptionMenu(frame, var, *self.well_names)
        
        select_sheet_button = Button(frame, 
                                     command=lambda: self.change_sheet_select(var),
                                     text="Select sheet")
        
        return dropdown, select_sheet_button
    
    def change_sheet_select(self, var):
        self.curr_sheet = self.get_well_by_name(var.get())
        print(self.curr_sheet)
    
class SheetsManager:
    def __init__(self, s1=None):
        self.well_data_path = s1
        self.worksheet_loaded = False
        
        # Workbooks (which haven't been opened yet)
        self.wb = None
        self.curr_sheet = None
        
        self.well_names = []

        
    def set_well_data_path(self, sheet):
        self.well_data_path = sheet
        
    def get_well_data_path(self):
        return self.well_data_path
    
    def select_well_data_popup(self, gui):
        self.set_well_data_path(fd.askopenfilename(filetypes=[("excel files", ".xlsx")]))
        gui.file_loaded = True 
        gui.check_file_loaded()
        self.generate_workbook()
        gui.setup_dropdown(gui.frames[1])
    
    # OpenPyXL functions
    def generate_workbook(self):
        self.wb = load_workbook(self.get_well_data_path())
        self.curr_sheet = self.wb.worksheets[1]
        # setup sheet names (so we can select all the different sheets)
        for i in self.wb.sheetnames[1:len(self.wb.sheetnames)-1]:
            self.well_names.append(i)
            
        print(self.well_names)
        #print(self.curr_sheet['G1'].value)
        
    def save_workbook(self, path):
        self.wb.save(path)
        
    def submit_entry(self, date, time, measure):
        
        date_val = parser.parse(date.get())
        time_val = parser.parse(time.get())
        measure_val = float(measure.get())
        
        next_row = self.get_next_empty_row_manual_table()
        self.insert_manual_data_into_row(date_val, time_val, measure_val, next_row)
        self.insert_formula_into_results_table(next_row)
        self.save_workbook(self.get_well_data_path())
        
        
        date.delete(0, END)
        time.delete(0, END)
        measure.delete(0, END)
        
    def get_next_empty_row_manual_table(self):
        for cell in self.curr_sheet['G']:
            if not cell.value:
                return cell.row
    
    def insert_manual_data_into_row(self, date, time, measure, row):
        self.curr_sheet[f'G{row}'].value = date
        self.curr_sheet[f'H{row}'].value = time
        self.curr_sheet[f'I{row}'].value = measure
    
    def insert_formula_into_results_table(self, row):
        self.curr_sheet[f'A{row}'].value = f"=G{row}"
        self.curr_sheet[f'B{row}'].value = f"=H{row}"
        self.curr_sheet[f'C{row}'].value = f"=D{row}/305"
        self.curr_sheet[f'C{row}'].number_format = '0.00'
        self.curr_sheet[f'D{row}'].value = f"=I{row}"